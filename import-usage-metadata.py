#!/usr/bin/env python3
"""
Import Historical Usage Data as Organization Metadata
Stores usage data in organizations.metadata JSONB field
"""

import openpyxl
from supabase import create_client
import json
from datetime import datetime

# Configuration
SUPABASE_URL = 'https://ksprdklquoskvjqsicvv.supabase.co'
SUPABASE_SERVICE_KEY = 'eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.eyJpc3MiOiJzdXBhYmFzZSIsInJlZiI6ImtzcHJka2xxdW9za3ZqcXNpY3Z2Iiwicm9sZSI6InNlcnZpY2Vfcm9sZSIsImlhdCI6MTc2MDE5NDcyMiwiZXhwIjoyMDc1NzcwNzIyfQ.VyRKDYIAzgbzmTZj29Lj5mdTvRGC5fetgPgfaOgCg54'

# Initialize Supabase client
supabase = create_client(SUPABASE_URL, SUPABASE_SERVICE_KEY)

# Load account mapping
with open('/home/ubuntu/b2bplus/distributor_accounts.json', 'r') as f:
    accounts = json.load(f)

account_lookup = {acc['name']: acc for acc in accounts}

# Load workbook
wb = openpyxl.load_workbook('/home/ubuntu/upload/BailyUsage.xlsx', read_only=False, data_only=True)

print("=" * 80)
print("IMPORTING HISTORICAL USAGE DATA TO ORGANIZATION METADATA")
print("=" * 80)
print()

total_imported = 0
customer_results = []

def clean_value(val):
    """Clean and convert value to appropriate type"""
    if val is None or val == '':
        return None
    if isinstance(val, (int, float)):
        return float(val) if val != 0 else None
    try:
        cleaned = str(val).strip().replace(',', '')
        return float(cleaned) if cleaned else None
    except:
        return None

def get_column_index(headers, possible_names):
    """Find column index by checking multiple possible names"""
    headers_str = [str(h) if h else '' for h in headers]
    headers_lower = [h.lower() for h in headers_str]
    
    for name in possible_names:
        name_lower = name.lower()
        if name_lower in headers_lower:
            return headers_lower.index(name_lower)
        for idx, h in enumerate(headers_lower):
            if name_lower in h or h in name_lower:
                return idx
    return None

# Process each customer
for sheet_name in wb.sheetnames:
    account = account_lookup.get(sheet_name)
    
    if not account:
        print(f"❌ {sheet_name}: Account not found")
        continue
    
    print(f"📊 {sheet_name}", end=" ")
    
    try:
        sheet = wb[sheet_name]
        headers = [cell.value for cell in sheet[1]]
        
        # Find columns
        item_col = get_column_index(headers, ['item#', 'item', 'item_number'])
        desc_col = get_column_index(headers, ['item description', 'description'])
        year_2022_col = get_column_index(headers, ['2022'])
        year_2023_col = get_column_index(headers, ['2023'])
        year_2024_col = get_column_index(headers, ['2024'])
        
        # Special cases - skip
        if sheet_name in ['Jims', 'Santos', 'D&S Dist', 'Biloxi Paper']:
            print("⚠️ Skipped")
            customer_results.append({'customer': sheet_name, 'status': 'skipped', 'imported': 0})
            continue
        
        # Bravos Jacksonville special handling
        start_row = 2
        if sheet_name == 'Bravos Jacksonville':
            first_row_vals = [cell.value for cell in sheet[1]]
            if first_row_vals and isinstance(first_row_vals[0], (int, float)):
                start_row = 1
                item_col = 0
                desc_col = 1
                year_2023_col = 3
        
        if item_col is None and desc_col is None:
            print("⚠️ Skipped (no columns)")
            customer_results.append({'customer': sheet_name, 'status': 'skipped', 'imported': 0})
            continue
        
        if year_2022_col is None and year_2023_col is None and year_2024_col is None:
            print("⚠️ Skipped (no usage)")
            customer_results.append({'customer': sheet_name, 'status': 'skipped', 'imported': 0})
            continue
        
        # Collect all usage data
        usage_data = []
        
        for row in sheet.iter_rows(min_row=start_row, values_only=True):
            if not any(row):
                break
            
            item_number = str(row[item_col]) if item_col is not None and len(row) > item_col and row[item_col] else None
            description = str(row[desc_col]) if desc_col is not None and len(row) > desc_col and row[desc_col] else None
            
            if item_number and item_number != 'None':
                try:
                    if '.' in item_number:
                        item_number = str(int(float(item_number)))
                except:
                    pass
            else:
                item_number = None
            
            usage_2022 = clean_value(row[year_2022_col]) if year_2022_col is not None and len(row) > year_2022_col else None
            usage_2023 = clean_value(row[year_2023_col]) if year_2023_col is not None and len(row) > year_2023_col else None
            usage_2024 = clean_value(row[year_2024_col]) if year_2024_col is not None and len(row) > year_2024_col else None
            
            if not item_number and not description:
                continue
            if not usage_2022 and not usage_2023 and not usage_2024:
                continue
            
            usage_data.append({
                'item_number': item_number,
                'description': description,
                'usage_2022': usage_2022,
                'usage_2023': usage_2023,
                'usage_2024': usage_2024,
            })
        
        if not usage_data:
            print("⚠️ No data")
            customer_results.append({'customer': sheet_name, 'status': 'skipped', 'imported': 0})
            continue
        
        # Update organization metadata
        metadata = {
            'historical_usage': {
                'source': 'bailey_usage_import',
                'imported_at': datetime.utcnow().isoformat(),
                'products': usage_data
            }
        }
        
        result = supabase.table('organizations').update({'metadata': metadata}).eq('id', account['id']).execute()
        
        if result.data:
            print(f"✅ {len(usage_data)} products")
            total_imported += len(usage_data)
            customer_results.append({'customer': sheet_name, 'status': 'success', 'imported': len(usage_data)})
        else:
            print("❌ Failed")
            customer_results.append({'customer': sheet_name, 'status': 'error', 'imported': 0})
    
    except Exception as e:
        print(f"❌ Error: {str(e)[:50]}")
        customer_results.append({'customer': sheet_name, 'status': 'error', 'imported': 0})

wb.close()

print()
print("=" * 80)
print("IMPORT COMPLETE")
print("=" * 80)
print()
print(f"✅ Total products imported: {total_imported}")
print(f"📊 Successful: {len([r for r in customer_results if r['status'] == 'success'])}/32")
print(f"⚠️  Skipped: {len([r for r in customer_results if r['status'] == 'skipped'])}/32")
print(f"❌ Errors: {len([r for r in customer_results if r['status'] == 'error'])}/32")
print()

# Save results
with open('/home/ubuntu/b2bplus/import_results.json', 'w') as f:
    json.dump({
        'total_imported': total_imported,
        'customers': customer_results,
        'timestamp': datetime.utcnow().isoformat()
    }, f, indent=2)

print("✅ Results saved to: /home/ubuntu/b2bplus/import_results.json")
