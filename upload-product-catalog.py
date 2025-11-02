#!/usr/bin/env python3
"""
Upload Metro Product Catalog to B2B+ Platform
Excludes: Baily Sale Price, Notes columns
"""

import openpyxl
from supabase import create_client
import json
from datetime import datetime

# Configuration
SUPABASE_URL = 'https://ksprdklquoskvjqsicvv.supabase.co'
SUPABASE_SERVICE_KEY = 'eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.eyJpc3MiOiJzdXBhYmFzZSIsInJlZiI6ImtzcHJka2xxdW9za3ZqcXNpY3Z2Iiwicm9sZSI6InNlcnZpY2Vfcm9sZSIsImlhdCI6MTc2MDE5NDcyMiwiZXhwIjoyMDc1NzcwNzIyfQ.VyRKDYIAzgbzmTZj29Lj5mdTvRGC5fetgPgfaOgCg54'

# Initialize Supabase client
supabase = create_client(SUPABASE_URL, SUPABASE_SERVICE_KEY)

# Load workbook
wb = openpyxl.load_workbook('/home/ubuntu/upload/MetroPricingComp(4).xlsx', read_only=False, data_only=True)
sheet = wb['Customer Copy']

print("=" * 80)
print("UPLOADING METRO PRODUCT CATALOG")
print("=" * 80)
print()

# Get headers
headers = [sheet.cell(1, col).value for col in range(1, sheet.max_column + 1)]

# Column mapping
col_map = {}
for idx, header in enumerate(headers, 1):
    if header:
        col_map[header.strip()] = idx

print("Columns found:")
for key in col_map.keys():
    print(f"  - {key}")
print()

# Process products
products = []
categories = set()
current_category = None

for row_idx in range(2, sheet.max_row + 1):
    item_num = sheet.cell(row_idx, col_map.get('Item #', 1)).value
    description = sheet.cell(row_idx, col_map.get('Item Description', 2)).value
    
    # Check if this is a category header (no item number, only description)
    if not item_num and description:
        current_category = description.strip()
        categories.add(current_category)
        continue
    
    # Skip empty rows
    if not item_num or not description:
        continue
    
    # Clean item number
    item_num = str(item_num).strip() if item_num else None
    
    # Get pricing data
    cost = sheet.cell(row_idx, col_map.get('Cost', 3)).value
    local_price = sheet.cell(row_idx, col_map.get('Local', 5)).value
    border_price = sheet.cell(row_idx, col_map.get('Bordering GA', 7)).value
    outer_price = sheet.cell(row_idx, col_map.get('Outer States', 10)).value
    pack_size = sheet.cell(row_idx, col_map.get('Pack Size', 8)).value
    
    # Clean and convert values
    def clean_number(val):
        if val is None or val == '':
            return None
        try:
            return float(val)
        except:
            return None
    
    cost = clean_number(cost)
    local_price = clean_number(local_price)
    border_price = clean_number(border_price)
    outer_price = clean_number(outer_price)
    
    # Clean pack size
    if pack_size:
        pack_size = str(pack_size).strip()
    
    # Create product record
    product = {
        'sku': item_num,
        'name': description.strip() if description else '',
        'description': description.strip() if description else '',
        'category': current_category,
        'pack_size': pack_size,
        'cost': cost,
        'price': local_price,  # Default price (Tier 1 - Local/Georgia)
        'stock': 1000,  # Default stock
        'is_active': True,
        'created_at': datetime.utcnow().isoformat(),
        'updated_at': datetime.utcnow().isoformat(),
    }
    
    products.append(product)

print(f"Total products to upload: {len(products)}")
print(f"Categories found: {len(categories)}")
print(f"Categories: {', '.join(sorted(categories))}")
print()

# Upload in batches
batch_size = 100
uploaded = 0
errors = 0

for i in range(0, len(products), batch_size):
    batch = products[i:i+batch_size]
    
    try:
        result = supabase.table('products').insert(batch).execute()
        uploaded += len(batch)
        print(f"✅ Uploaded batch {i//batch_size + 1}: {len(batch)} products")
    except Exception as e:
        print(f"⚠️  Batch {i//batch_size + 1} error: {str(e)[:100]}")
        # Try one by one
        for product in batch:
            try:
                supabase.table('products').insert(product).execute()
                uploaded += 1
            except Exception as e2:
                errors += 1
                print(f"  ❌ Failed: {product['sku']} - {str(e2)[:50]}")

wb.close()

print()
print("=" * 80)
print("UPLOAD COMPLETE")
print("=" * 80)
print()
print(f"✅ Successfully uploaded: {uploaded} products")
print(f"❌ Errors: {errors}")
print()

# Save summary
summary = {
    'total_products': len(products),
    'uploaded': uploaded,
    'errors': errors,
    'categories': list(categories),
    'timestamp': datetime.utcnow().isoformat()
}

with open('/home/ubuntu/b2bplus/product_upload_results.json', 'w') as f:
    json.dump(summary, f, indent=2)

print("Results saved to: /home/ubuntu/b2bplus/product_upload_results.json")
