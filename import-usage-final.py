#!/usr/bin/env python3
"""
Import Historical Usage Data from BailyUsage.xlsx
Handles all 32 customers including special cases
"""

import openpyxl
from supabase import create_client
import json
from datetime import datetime

# Configuration
SUPABASE_URL = 'https://ksprdklquoskvjqsicvv.supabase.co'
SUPABASE_SERVICE_KEY = 'eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.eyJpc3MiOiJzdXBhYmFzZSIsInJlZiI6ImtzcHJka2xxdW9za3ZqcXNpY3Z2Iiwicm9sZSI6InNlcnZpY2Vfcm9sZSIsImlhdCI6MTc2MDE5NDcyMiwiZXhwIjoyMDc1NzcwNzIyfQ.VyRKDYIAzgbzmTZj29Lj5mdTvRGC5fetgPgfaOgCg54'

# Initialize Supabase client
supabase = create_client(SUPABASE_URL, SUPABASE_SERVICE_KEY)

# Load account mapping
with open('/home/ubuntu/b2bplus/distributor_accounts.json', 'r') as f:
    accounts = json.load(f)

# Create lookup dict
account_lookup = {acc['name']: acc for acc in accounts}

# Load workbook
wb = openpyxl.load_workbook('/home/ubuntu/upload/BailyUsage.xlsx', read_only=False, data_only=True)

print("=" * 80)
print("IMPORTING HISTORICAL USAGE DATA")
print("=" * 80)
print()

total_imported = 0
total_skipped = 0
customer_results = []

def clean_value(val):
    """Clean and convert value to appropriate type"""
    if val is None or val == '':
        return None
    if isinstance(val, (int, float)):
        return float(val) if val != 0 else None
    try:
        cleaned = str(val).strip().replace(',', '')
        return float(cleaned) if cleaned else None
    except:
        return None

def get_column_index(headers, possible_names):
    """Find column index by checking multiple possible names"""
    headers_str = [str(h) if h else '' for h in headers]
    headers_lower = [h.lower() for h in headers_str]
    
    for name in possible_names:
        name_lower = name.lower()
        # Exact match
        if name_lower in headers_lower:
            return headers_lower.index(name_lower)
        # Partial match
        for idx, h in enumerate(headers_lower):
            if name_lower in h or h in name_lower:
                return idx
    return None

# Process each customer
for sheet_name in wb.sheetnames:
    account = account_lookup.get(sheet_name)
    
    if not account:
        print(f"❌ {sheet_name}: Account not found")
        continue
    
    print(f"📊 {sheet_name}")
    
    try:
        sheet = wb[sheet_name]
        
        # Get headers
        headers = [cell.value for cell in sheet[1]]
        
        # Find columns
        item_col = get_column_index(headers, ['item#', 'item', 'item_number', 'itemnumber'])
        desc_col = get_column_index(headers, ['item description', 'description', 'item_description'])
        year_2022_col = get_column_index(headers, ['2022', '2022.0'])
        year_2023_col = get_column_index(headers, ['2023', '2023.0'])
        year_2024_col = get_column_index(headers, ['2024', '2024.0'])
        category_col = get_column_index(headers, ['category', 'cat'])
        
        # Special cases
        if sheet_name == 'Jims':
            print(f"   ⚠️  Skipped: No product information")
            customer_results.append({'customer': sheet_name, 'status': 'skipped', 'reason': 'No product info', 'imported': 0})
            continue
        
        if sheet_name == 'Santos':
            print(f"   ⚠️  Skipped: No usage quantities")
            customer_results.append({'customer': sheet_name, 'status': 'skipped', 'reason': 'No quantities', 'imported': 0})
            continue
        
        if sheet_name in ['D&S Dist', 'Biloxi Paper']:
            # Check if actually empty
            has_data = False
            for row in sheet.iter_rows(min_row=2, max_row=5, values_only=True):
                if any(row):
                    has_data = True
                    break
            if not has_data:
                print(f"   ⚠️  Skipped: Empty sheet")
                customer_results.append({'customer': sheet_name, 'status': 'skipped', 'reason': 'Empty', 'imported': 0})
                continue
        
        # Bravos Jacksonville special handling
        start_row = 2
        if sheet_name == 'Bravos Jacksonville':
            # Check if first row is actually data
            first_row_vals = [cell.value for cell in sheet[1]]
            if first_row_vals and isinstance(first_row_vals[0], (int, float)):
                start_row = 1
                item_col = 0
                desc_col = 1
                year_2023_col = 3
        
        # Validate we have minimum columns
        if item_col is None and desc_col is None:
            print(f"   ⚠️  Skipped: Cannot identify columns")
            customer_results.append({'customer': sheet_name, 'status': 'skipped', 'reason': 'No columns', 'imported': 0})
            continue
        
        if year_2022_col is None and year_2023_col is None and year_2024_col is None:
            print(f"   ⚠️  Skipped: No usage columns")
            customer_results.append({'customer': sheet_name, 'status': 'skipped', 'reason': 'No usage data', 'imported': 0})
            continue
        
        # Import data
        imported_count = 0
        batch = []
        
        for row in sheet.iter_rows(min_row=start_row, values_only=True):
            # Check if empty
            if not any(row):
                break
            
            # Extract values
            item_number = str(row[item_col]) if item_col is not None and len(row) > item_col and row[item_col] else None
            description = str(row[desc_col]) if desc_col is not None and len(row) > desc_col and row[desc_col] else None
            category = str(row[category_col]) if category_col is not None and len(row) > category_col and row[category_col] else None
            
            # Clean item number
            if item_number and item_number != 'None':
                try:
                    if '.' in item_number:
                        item_number = str(int(float(item_number)))
                except:
                    pass
            else:
                item_number = None
            
            # Get usage values
            usage_2022 = clean_value(row[year_2022_col]) if year_2022_col is not None and len(row) > year_2022_col else None
            usage_2023 = clean_value(row[year_2023_col]) if year_2023_col is not None and len(row) > year_2023_col else None
            usage_2024 = clean_value(row[year_2024_col]) if year_2024_col is not None and len(row) > year_2024_col else None
            
            # Skip if no data
            if not item_number and not description:
                continue
            if not usage_2022 and not usage_2023 and not usage_2024:
                continue
            
            # Add to batch
            batch.append({
                'organization_id': account['id'],
                'item_number': item_number,
                'item_description': description,
                'usage_2022': usage_2022,
                'usage_2023': usage_2023,
                'usage_2024': usage_2024,
                'category': category,
                'source': 'bailey_usage_import',
                'created_at': datetime.utcnow().isoformat(),
                'updated_at': datetime.utcnow().isoformat(),
            })
            
            # Insert in batches of 100
            if len(batch) >= 100:
                try:
                    supabase.table('customer_annual_usage').insert(batch).execute()
                    imported_count += len(batch)
                    batch = []
                except Exception as e:
                    print(f"   ⚠️  Batch insert error: {e}")
                    # Try one by one
                    for item in batch:
                        try:
                            supabase.table('customer_annual_usage').insert(item).execute()
                            imported_count += 1
                        except:
                            pass
                    batch = []
        
        # Insert remaining
        if batch:
            try:
                supabase.table('customer_annual_usage').insert(batch).execute()
                imported_count += len(batch)
            except Exception as e:
                print(f"   ⚠️  Final batch error: {e}")
                for item in batch:
                    try:
                        supabase.table('customer_annual_usage').insert(item).execute()
                        imported_count += 1
                    except:
                        pass
        
        print(f"   ✅ Imported: {imported_count} products")
        total_imported += imported_count
        
        customer_results.append({
            'customer': sheet_name,
            'status': 'success',
            'imported': imported_count
        })
    
    except Exception as e:
        print(f"   ❌ Error: {str(e)}")
        customer_results.append({
            'customer': sheet_name,
            'status': 'error',
            'message': str(e),
            'imported': 0
        })
        total_skipped += 1

wb.close()

print()
print("=" * 80)
print("IMPORT COMPLETE")
print("=" * 80)
print()
print(f"✅ Total products imported: {total_imported}")
print(f"📊 Customers processed: {len([r for r in customer_results if r['status'] == 'success'])}")
print(f"⚠️  Customers skipped: {len([r for r in customer_results if r['status'] == 'skipped'])}")
print(f"❌ Customers with errors: {len([r for r in customer_results if r['status'] == 'error'])}")
print()

# Print summary by customer
print("Customer Summary:")
for result in customer_results:
    status_icon = "✅" if result['status'] == 'success' else "⚠️" if result['status'] == 'skipped' else "❌"
    print(f"  {status_icon} {result['customer']}: {result['imported']} products")

# Save results
with open('/home/ubuntu/b2bplus/import_results.json', 'w') as f:
    json.dump({
        'total_imported': total_imported,
        'total_skipped': total_skipped,
        'customers': customer_results,
        'timestamp': datetime.utcnow().isoformat()
    }, f, indent=2)

print()
print("Results saved to: /home/ubuntu/b2bplus/import_results.json")
